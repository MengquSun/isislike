"""Custom export: molecule search + XLSX/CSV generation with pivoted database fields."""

from __future__ import annotations

import csv
import re
import zipfile
from io import BytesIO, StringIO
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

from app.config import settings
from app.models.schemas import EXPORT_ATTRIBUTES
from app.services import database_client, rdkit_service, supabase_client
from app.services.rdkit_service import RDKitError

ATTRIBUTE_LABELS: dict[str, str] = {
    "canonical_smiles": "Canonical SMILES",
    "molecular_weight": "Molecular weight",
    "molecular_formula": "Formula",
    "created_at": "Created",
    "updated_at": "Updated",
    "name": "Name",
    "notes": "Notes",
    "structure_image": "2D structure",
}

_STRUCTURE_IMAGE_WIDTH = 200
_STRUCTURE_IMAGE_HEIGHT = 150
_STRUCTURE_ROW_HEIGHT = 115

_ILLEGAL_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")
_ILLEGAL_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def _format_date(value: str | None) -> str:
    if not value:
        return ""
    return value[:10] if len(value) >= 10 else value


def _attribute_value(molecule: dict[str, Any], attr: str) -> str:
    if attr == "created_at":
        return _format_date(molecule.get("created_at"))
    if attr == "updated_at":
        return _format_date(molecule.get("updated_at"))
    if attr == "molecular_weight":
        mw = molecule.get("molecular_weight")
        return "" if mw is None else str(mw)
    val = molecule.get(attr)
    return "" if val is None else str(val)


def _record_value_display(value: dict[str, Any]) -> str:
    if value.get("number_value") is not None:
        return str(value["number_value"])
    if value.get("date_value"):
        return str(value["date_value"])
    if value.get("text_value"):
        return str(value["text_value"])
    return ""


def _sanitize_sheet_name(base: str, used: set[str]) -> str:
    name = _ILLEGAL_SHEET_CHARS.sub("_", base).strip() or "Sheet"
    if len(name) > 31:
        name = name[:31]
    candidate = name
    n = 2
    while candidate in used:
        suffix = f"_{n}"
        candidate = name[: 31 - len(suffix)] + suffix
        n += 1
    used.add(candidate)
    return candidate


def _sanitize_filename(base: str, used: set[str]) -> str:
    name = _ILLEGAL_FILENAME_CHARS.sub("_", base).strip() or "chemical"
    if len(name) > 80:
        name = name[:80]
    candidate = name
    n = 2
    while candidate in used:
        suffix = f"_{n}"
        candidate = name[: 80 - len(suffix)] + suffix
        n += 1
    used.add(candidate)
    return candidate


async def _field_sort_map() -> dict[str, int]:
    defs = await database_client.list_all_field_definitions()
    result: dict[str, int] = {}
    for fd in defs:
        name = fd.get("name")
        if not name:
            continue
        order = fd.get("sort_order", 0)
        if name not in result or order < result[name]:
            result[name] = order
    return result


def _order_field_columns(
    field_names: set[str],
    *,
    field_sort: str,
    sort_map: dict[str, int],
) -> list[str]:
    if field_sort == "definition_order":
        return sorted(field_names, key=lambda n: (sort_map.get(n, 9999), n.lower()))
    return sorted(field_names, key=lambda n: n.lower())


def _pivot_records(
    records: list[dict[str, Any]],
    *,
    field_sort: str = "alphabetical",
    sort_map: dict[str, int] | None = None,
) -> tuple[list[str], list[dict[str, str]]]:
    """Group field values by record_id; merge columns by field_name."""
    field_names: set[str] = set()
    rows: list[dict[str, str]] = []

    for record in records:
        row: dict[str, str] = {}
        for value in record.get("values") or []:
            field_name = value.get("field_name")
            if not field_name:
                continue
            field_names.add(field_name)
            row[field_name] = _record_value_display(value)
        rows.append(row)

    columns = _order_field_columns(
        field_names,
        field_sort=field_sort,
        sort_map=sort_map or {},
    )
    return columns, rows


def _structure_png_bytes(molecule: dict[str, Any]) -> bytes | None:
    smiles = molecule.get("canonical_smiles")
    if not smiles:
        return None
    return rdkit_service.structure_png_from_smiles(smiles)


def _attribute_rows(
    molecule: dict[str, Any],
    attributes: list[str],
) -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = []
    for attr in attributes:
        if attr not in EXPORT_ATTRIBUTES or attr == "structure_image":
            continue
        label = ATTRIBUTE_LABELS.get(attr, attr)
        rows.append((label, _attribute_value(molecule, attr)))
    return rows


async def search_for_preview(
    *,
    canonical_smiles: str | None = None,
    formula: str | None = None,
    name: str | None = None,
    all_chemicals: bool = False,
) -> list[dict[str, Any]]:
    limit = settings.export_max_molecules

    if all_chemicals:
        return await supabase_client.list_molecules(limit=limit)

    smiles_filter = canonical_smiles
    if canonical_smiles and canonical_smiles.strip():
        try:
            props = rdkit_service.canonicalize_smiles(canonical_smiles.strip())
            smiles_filter = props.canonical_smiles
        except RDKitError:
            return []

    formula_filter = formula.strip() if formula and formula.strip() else None
    name_filter = name.strip() if name and name.strip() else None

    if not smiles_filter and not formula_filter and not name_filter:
        return await supabase_client.list_molecules(limit=limit)

    return await supabase_client.search_molecules(
        canonical_smiles=smiles_filter,
        molecular_formula=formula_filter,
        name=name_filter,
        limit=limit,
    )


def _write_molecule_sheet(
    ws,
    molecule: dict[str, Any],
    attributes: list[str],
    records: list[dict[str, Any]],
    *,
    field_sort: str,
    sort_map: dict[str, int],
) -> None:
    row_idx = 1
    ws.cell(row=row_idx, column=1, value="Attribute")
    ws.cell(row=row_idx, column=2, value="Value")
    row_idx += 1

    for attr in attributes:
        if attr not in EXPORT_ATTRIBUTES:
            continue
        label = ATTRIBUTE_LABELS.get(attr, attr)
        if attr == "structure_image":
            ws.cell(row=row_idx, column=1, value=label)
            png = _structure_png_bytes(molecule)
            if png:
                img = XLImage(BytesIO(png))
                img.width = _STRUCTURE_IMAGE_WIDTH
                img.height = _STRUCTURE_IMAGE_HEIGHT
                ws.add_image(img, f"B{row_idx}")
                ws.row_dimensions[row_idx].height = _STRUCTURE_ROW_HEIGHT
            else:
                ws.cell(row=row_idx, column=2, value="")
            row_idx += 1
            continue

        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=_attribute_value(molecule, attr))
        row_idx += 1

    if records:
        row_idx += 1
        columns, pivot_rows = _pivot_records(
            records,
            field_sort=field_sort,
            sort_map=sort_map,
        )
        if columns:
            for col_idx, field_name in enumerate(columns, start=1):
                ws.cell(row=row_idx, column=col_idx, value=field_name)
            row_idx += 1
            for pivot_row in pivot_rows:
                for col_idx, field_name in enumerate(columns, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=pivot_row.get(field_name, ""))
                row_idx += 1


def _molecule_to_csv(
    molecule: dict[str, Any],
    attributes: list[str],
    records: list[dict[str, Any]],
    *,
    field_sort: str,
    sort_map: dict[str, int],
    structure_filename: str | None = None,
) -> str:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Attribute", "Value"])
    for attr in attributes:
        if attr not in EXPORT_ATTRIBUTES:
            continue
        label = ATTRIBUTE_LABELS.get(attr, attr)
        if attr == "structure_image":
            writer.writerow([label, structure_filename or ""])
            continue
        writer.writerow([label, _attribute_value(molecule, attr)])

    if records:
        writer.writerow([])
        columns, pivot_rows = _pivot_records(
            records,
            field_sort=field_sort,
            sort_map=sort_map,
        )
        if columns:
            writer.writerow(columns)
            for pivot_row in pivot_rows:
                writer.writerow([pivot_row.get(col, "") for col in columns])

    return buffer.getvalue()


async def build_custom_export(
    molecule_ids: list[str],
    attributes: list[str],
    *,
    fmt: str = "xlsx",
    field_sort: str = "alphabetical",
) -> tuple[bytes, str, str]:
    """Return (content, media_type, filename)."""
    valid_attrs = [a for a in attributes if a in EXPORT_ATTRIBUTES]
    include_structure = "structure_image" in valid_attrs
    molecules = await supabase_client.fetch_molecules_by_ids(
        molecule_ids,
        include_structure=include_structure,
    )
    if not molecules:
        raise ValueError("No molecules found for export")

    records_by_molecule = await database_client.fetch_records_by_molecule_ids(molecule_ids)
    sort_map = await _field_sort_map() if field_sort == "definition_order" else {}

    if fmt == "csv":
        zip_buffer = BytesIO()
        used_names: set[str] = set()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for i, molecule in enumerate(molecules):
                base_name = (molecule.get("name") or "").strip() or f"Chemical {i + 1}"
                safe_base = _sanitize_filename(base_name, used_names)
                filename = safe_base + ".csv"
                png = _structure_png_bytes(molecule) if include_structure else None
                structure_filename = f"{safe_base}_structure.png" if png else None
                content = _molecule_to_csv(
                    molecule,
                    valid_attrs,
                    records_by_molecule.get(molecule["id"], []),
                    field_sort=field_sort,
                    sort_map=sort_map,
                    structure_filename=structure_filename,
                )
                zf.writestr(filename, content)
                if png and structure_filename:
                    zf.writestr(structure_filename, png)
        return (
            zip_buffer.getvalue(),
            "application/zip",
            "custom_export.zip",
        )

    wb = Workbook()
    wb.remove(wb.active)
    used_names: set[str] = set()

    for i, molecule in enumerate(molecules):
        base_name = (molecule.get("name") or "").strip() or f"Chemical {i + 1}"
        sheet_name = _sanitize_sheet_name(base_name, used_names)
        ws = wb.create_sheet(title=sheet_name)
        _write_molecule_sheet(
            ws,
            molecule,
            valid_attrs,
            records_by_molecule.get(molecule["id"], []),
            field_sort=field_sort,
            sort_map=sort_map,
        )

    buffer = BytesIO()
    wb.save(buffer)
    return (
        buffer.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "custom_export.xlsx",
    )
